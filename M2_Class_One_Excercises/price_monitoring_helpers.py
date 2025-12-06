import requests
from bs4 import BeautifulSoup

def load_website():
   website = requests.get('https://web-scraping.dev/products')
   website.raise_for_status()

   soup = BeautifulSoup(website.text, 'html.parser')

  #  print(soup.prettify())

def load_excel_file():
  from openpyxl import load_workbook

  # Get the workbook
  price_info_sheet = load_workbook('C:/OKCoders_Pro_Python/sales_data.xlsx')
  active_sheet = price_info_sheet.active

  # Store data as a list of dictionaries
  data = []

  # Skipping the header row (since we do not need that info), and loading data beginning at row 2
  for row in active_sheet.iter_rows(min_row = 2, values_only = True):
      product, quantity_sold, unit_price, total_sales = row

      item = {
          'product': product.strip().lower(),
          'quantity_sold': quantity_sold,
          'unit_price': unit_price,
          'total_sales': total_sales
      }

      data.append(item)

  # Checking for now (will remove later)
  for item in data:
      print(item)
